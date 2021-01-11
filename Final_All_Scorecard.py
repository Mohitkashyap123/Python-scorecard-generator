import openpyxl
from reportlab.pdfgen import canvas
from reportlab.lib import utils
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from PIL import Image
from reportlab.platypus import Table,TableStyle
from reportlab.lib import colors

import numpy as np 
import matplotlib.pyplot as plt

# Reading Excel File
wb = openpyxl.load_workbook('data.xlsx')
sheet = wb.get_sheet_by_name('Raw data')
print(f"{sheet['C4'].value}")

total_student = 20
avg_rows = 5 # Number of questions
c = sheet.max_column
# PERCENTILE CALCULATION
per = []
for i in range(3,99,5) :
    t = 0
    for j in range(5) :
        #print(i+j,"score ",sheet.cell(row=(i+j), column=20).value)
        #print(i+j)
        sc = int(sheet.cell(row=(i+j), column=20).value)
        t += sc
    per.append(t)
per.sort()
print(f"SCORE DONE \nPERCENTILE DONE")
r = 1
for k in range(3,(total_student*avg_rows),avg_rows) :
    # PERSONAL INFORMATION
    name = sheet.cell(row=k, column=2).value
    grade = sheet.cell(row=k, column=4).value
    school = sheet.cell(row=k, column=6).value
    city = sheet.cell(row=k, column=8).value
    country = sheet.cell(row=k, column=10).value

    registration = sheet.cell(row=k, column=3).value
    gender = sheet.cell(row=k, column=5).value
    date_of_birth = sheet.cell(row=k, column=7).value
    date_of_test = sheet.cell(row=k, column=9).value
    extra_time = sheet.cell(row=k, column=11).value

    #Generating Pdf
    fileName = f'{r}.pdf'
    pdf = canvas.Canvas(fileName)
    pdf.setPageSize((700, 1150))
    top_x = 700
    left_y = 1150

    pdf.setTitle(f"{name}")
    # Title
    pdf.setFont('Helvetica-Bold',20)
    pdf.drawCentredString(350, left_y-30, "Wisdom Tests and Math Challenge")
    pdf.setFont('Helvetica',12)

    # LOGO
    from PIL import Image
    logo = Image.open('logo.jpg')
    width ,height = logo.size
    ratio =  width/height
    w = 80
    h = w/ratio
    pic = Image.open(f"Pics/{(r//2)+1}.jpg")
    r += 1
    width1 ,height1 = pic.size
    ratio1 =  width1/height1
    w1 = 100
    h1 = w1/ratio1
    pdf.drawInlineImage(logo, 250, left_y-100,w,h)
    pdf.drawInlineImage(pic, 550, left_y-120,90,100)

    # Adding Personal Data
    pdf.setFont('Helvetica-Bold',12)
    pdf.drawString(40, left_y-140, 'Name of candidate : ')
    pdf.drawString(40, left_y-160, 'Grade : ')
    pdf.drawString(40, left_y-180, 'School : ')
    pdf.drawString(40, left_y-200, 'City of Residence : ')
    pdf.drawString(40, left_y-220, 'Country : ')

    pdf.setFont('Helvetica',12)
    pdf.drawString(175, left_y-140, name)
    pdf.drawString(175, left_y-160, str(grade))
    pdf.drawString(175, left_y-180, school)
    pdf.drawString(175, left_y-200, city)
    pdf.drawString(175, left_y-220, country)

    pdf.setFont('Helvetica-Bold',12)
    pdf.drawString(360, left_y-140, 'Registration No. : ')
    pdf.drawString(360, left_y-160, 'Gender : ')
    pdf.drawString(360, left_y-180, 'Date of Test : ')
    pdf.drawString(360, left_y-200, 'Date of Birth : ')
    pdf.drawString(360, left_y-220, 'Extra Time Assistance : ')

    pdf.setFont('Helvetica',12)
    pdf.drawString(505, left_y-140, str(registration))
    pdf.drawString(505, left_y-160, gender)
    pdf.drawString(505, left_y-180, str(date_of_test))
    pdf.drawString(505, left_y-200, str(date_of_birth))
    pdf.drawString(505, left_y-220, extra_time)

    # TEST RESULT
    data = []
    heading = ["Question No. " ,"Time spent on\nquestion (sec)" ," Score if\n correct " ," Score if\n incorrect " ," Attempt status " ," what you\n  marked " ,"Correct\nAnswer" ,"Outcome" ,"Your score"]
    data.append(heading)
    total_score = 0
    for i in range(1,6) :
        Q_num = "    " + str(sheet.cell(row=k+i-1, column=12).value)
        time = "    " + str(sheet.cell(row=k+i-1, column=13).value)
        correct_score = "    " + str(sheet.cell(row=k+i-1, column=14).value)
        incorrect_score = "    " + str(sheet.cell(row=k+i-1, column=15).value)
        status = sheet.cell(row=k+i-1, column=16).value
        your_ans = "    " + str(sheet.cell(row=k+i-1, column=17).value)
        correct_ans = "    " + str(sheet.cell(row=k+i-1, column=18).value)
        outcome = sheet.cell(row=k+i-1, column=19).value
        your_score = "    " + str(sheet.cell(row=k+i-1, column=20).value)
        ques_data = [Q_num ,time ,correct_score ,incorrect_score ,status ,your_ans ,correct_ans ,outcome ,your_score]
        data.append(ques_data)
        total_score += int(sheet.cell(row=k+i-1, column=20).value)

    grid = [('GRID', (0,0), (-1,-1), 0.25, colors.black), ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold')]
    t=Table(data, repeatRows=1, style=TableStyle(grid))
    t.wrapOn(pdf,400 ,500 )
    t.drawOn(pdf, 40, 720)

    # Total Score
    pdf.setFont('Helvetica-Bold',16)
    pdf.drawString(40, 680, 'Total Score : ')
    pdf.setFont('Helvetica',16)
    pdf.drawString(150, 680, str(total_score))

    # PERCENTILE
    pdf.setFont('Helvetica-Bold',16)
    prt = per.index(total_score)/total_student
    pdf.drawString(50, 650, 'Overall Percentile : ')
    pdf.setFont('Helvetica',16)
    pdf.drawString(220,650, str(prt*100)+'%')
    
    #import Image
    from io import BytesIO
    from reportlab.lib.units import inch, cm
    from reportlab.lib.utils import ImageReader
    # Bar Graph
    x = []
    y = []
    status = []
    outcome = []
    for i in range(1,6) :
        Q_num = str(sheet.cell(row=k+i-1, column=12).value)
        print(k+i-1)
        time = int(sheet.cell(row=k+i-1, column=13).value)
        status.append(str(sheet.cell(row=k+i-1, column=16).value))
        outcome.append(str(sheet.cell(row=k+i-1, column=19).value))
    
        x.append(Q_num)
        y.append(time)
    fig = plt.figure(figsize = (6, 5))
    barplot = plt.bar(x,y,width=0.3, align='center',capsize=8)
    for bar in barplot:
        yval = bar.get_height()
        plt.text(bar.get_x() + bar.get_width()/2.0, yval, int(yval), va='bottom')

    plt.xlabel("Question number") 
    plt.ylabel("Time (sec) ") 
    plt.title("Time",fontsize=20) 

    plt.rc('xtick',labelsize=16)
    plt.rc('ytick',labelsize=16)

    imgdata = BytesIO()
    fig.savefig(imgdata, format='png')
    imgdata.seek(0)  # rewind the data

    Image = ImageReader(imgdata)
    w = 250
    h = 180
    pdf.drawImage(Image, 50, 450, w, h)
    plt.close(fig=None)
    # Pie Chart Time question
    time = y
    t = sum(y)
    explode = (0, 0.1, 0, 0,0)

    fig1, ax1 = plt.subplots()
    plt.title("Time Spent as function\n   of Total Time",fontsize=18)
    ax1.pie(time, explode=explode, labels=x, autopct='%1.1f%%',
        shadow=True, startangle=90)
    ax1.axis('equal') 

    imgdata = BytesIO()
    fig1.savefig(imgdata, format='png')
    imgdata.seek(0)  # rewind the data
    Image = ImageReader(imgdata)
    w = 290
    h = 230
    pdf.drawImage(Image, 350, 410, w, h)
    plt.close(fig=None)
    # Attempted Unattempted Pie Chart
    count = 0
    for i in status :
        if i == 'Attempted' :
            count += 1
    x = [count,avg_rows-count]
    y = ['Attempted','Unattempted']
    fig2, ax2 = plt.subplots()
    plt.title("Attempts",fontsize=20)
    ax2.axis('equal')
    explode = (0, 0.1)
    if count == 0:
        x = [avg_rows-count]
        y = ['Unattempted']
        ax2.pie(x, labels=y, autopct='%1.1f%%',
        shadow=True, startangle=90)
    elif avg_rows-count == 0 :
        x = [count]
        y = ['attempted']
        ax2.pie(x, labels=y, autopct='%1.1f%%',
        shadow=True, startangle=90)
    else :
        ax2.pie(x, explode=explode, labels=y, autopct='%1.1f%%',
        shadow=True, startangle=90)

    imgdata = BytesIO()
    fig2.savefig(imgdata, format='png')
    imgdata.seek(0)  # rewind the data

    Image = ImageReader(imgdata)
    w = 290
    h = 230
    pdf.drawImage(Image, 0, 200, w, h)
    plt.close(fig=None)
    # Correct Incorrect pie Chart
    count1 = 0
    count2 = 0
    for i in outcome :
        if i == 'Correct' :
            count1 += 1
        elif i == 'Incorrect' :
            count2 += 1
    x = [count1,count2]
    y = ['Correct','Incorrect']
    fig3, ax3 = plt.subplots()
    plt.title("Accuracy from Attempted Questions",fontsize=20)
    ax3.axis('equal')
    explode = (0, 0.1)
    if count2 == 0:
        x = [count1]
        y = ['Correct']
        ax3.pie(x, labels=y, autopct='%1.1f%%',
        shadow=True, startangle=90)
    elif count1 == 0 :
        x = [count2]
        y = ['Incorrect']
        ax3.pie(x, labels=y, autopct='%1.1f%%',
        shadow=True, startangle=90)
    else :
        ax3.pie(x, explode=explode, labels=y, autopct='%1.1f%%',
        shadow=True, startangle=90)
    
    imgdata = BytesIO()
    fig3.savefig(imgdata, format='png')
    imgdata.seek(0)  # rewind the data

    Image = ImageReader(imgdata)
    w = 290
    h = 230
    pdf.drawImage(Image, 370, 200, w, h)
    plt.close(fig=None)
    # Correct Incorrect Unattempted Pie chart
    x = [avg_rows-count,count1,count2]
    y = ['Unattempted','Correct','Incorrect']
    
    fig4, ax4 = plt.subplots()
    explode = (0, 0.1,0)
    if avg_rows == count :
        explode = (0, 0.1)
        x = [count1,count2]
        y = ['Correct','Incorrect']
    elif count1 == 0 :
        explode = (0, 0.1)
        x = [avg_rows-count,count2]
        y = ['Unattempted','Incorrect']
    elif count2 == 0 :
        explode = (0, 0.1)
        x = [avg_rows-count,count1]
        y = ['Unattempted','Correct']
    ax4.pie(x, explode=explode, labels=y, autopct='%1.1f%%',
            shadow=True, startangle=90)
    ax4.axis('equal')
    plt.title("Overall performance In Test",fontsize=20)

    imgdata = BytesIO()
    fig4.savefig(imgdata, format='png')
    imgdata.seek(0)  # rewind the data

    Image = ImageReader(imgdata)
    w = 290
    h = 230
    pdf.drawImage(Image, 180, -10, w, h)
    plt.close(fig=None)
    pdf.showPage()
    pdf.save()